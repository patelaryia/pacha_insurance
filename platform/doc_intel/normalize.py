"""Deterministic document normalisation and text-layer extraction."""

from __future__ import annotations

import csv
import io
import json
import re
import tempfile
from dataclasses import dataclass
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import Any, Protocol

import fitz
import pytesseract
from openpyxl import load_workbook
from PIL import Image

from claim_core import BlobStore


class NormaliseError(RuntimeError):
    """The immutable source could not be converted into a readable PDF."""


class OcrEngine(Protocol):
    def words(self, page_png_bytes: bytes) -> list[dict[str, Any]]: ...


class TesseractOcrEngine:
    """Tesseract 5 adapter returning normalized word boxes."""

    def words(self, page_png_bytes: bytes) -> list[dict[str, Any]]:
        image = Image.open(io.BytesIO(page_png_bytes))
        width, height = image.size
        data = pytesseract.image_to_data(
            image, output_type=pytesseract.Output.DICT, config="--psm 6"
        )
        words = []
        for index, raw_text in enumerate(data["text"]):
            text = raw_text.strip()
            if not text:
                continue
            left = int(data["left"][index])
            top = int(data["top"][index])
            right = left + int(data["width"][index])
            bottom = top + int(data["height"][index])
            words.append(
                {
                    "text": text,
                    "bbox": [
                        left / width,
                        top / height,
                        right / width,
                        bottom / height,
                    ],
                }
            )
        return words


@dataclass(frozen=True)
class NormaliseResult:
    pdf_key: str
    page_count: int
    text_keys: list[str]
    page_keys: list[str]
    first_text: str
    page_text_coverages: list[float]
    page_has_native_text: list[bool]
    email_subject: str | None


def _plain_text_pdf(text: str) -> bytes:
    document = fitz.open()
    page = document.new_page()
    y = 54.0
    for raw_line in text.splitlines() or [""]:
        chunks = [raw_line[index : index + 95] for index in range(0, len(raw_line), 95)] or [""]
        for line in chunks:
            if y > page.rect.height - 54:
                page = document.new_page()
                y = 54.0
            page.insert_text((54, y), line, fontsize=10)
            y += 14
    return document.tobytes()


def _email_text(content: bytes) -> str:
    message = BytesParser(policy=policy.default).parsebytes(content)
    body = message.get_body(preferencelist=("plain",))
    if body is not None:
        return body.get_content()
    if message.is_multipart():
        parts = [
            part.get_content()
            for part in message.walk()
            if part.get_content_type() == "text/plain"
        ]
        return "\n".join(map(str, parts))
    return str(message.get_content())


def _email_subject(content: bytes) -> str | None:
    subject = BytesParser(policy=policy.default).parsebytes(content).get("subject")
    return str(subject) if subject is not None else None


def _msg_text(content: bytes) -> str:
    import extract_msg

    with tempfile.NamedTemporaryFile(suffix=".msg") as source:
        source.write(content)
        source.flush()
        message = extract_msg.Message(source.name)
        try:
            return message.body or ""
        finally:
            message.close()


def _msg_subject(content: bytes) -> str | None:
    import extract_msg

    with tempfile.NamedTemporaryFile(suffix=".msg") as source:
        source.write(content)
        source.flush()
        message = extract_msg.Message(source.name)
        try:
            return message.subject or None
        finally:
            message.close()


def _image_pdf(content: bytes) -> bytes:
    with Image.open(io.BytesIO(content)) as source:
        image = source.convert("RGB")
        output = io.BytesIO()
        image.save(output, format="PDF", resolution=300)
        return output.getvalue()


def _xlsx_pdf_and_snapshots(
    content: bytes, document_id: str, blob_store: BlobStore
) -> bytes:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    combined: list[str] = []
    try:
        for position, worksheet in enumerate(workbook.worksheets, start=1):
            output = io.StringIO(newline="")
            writer = csv.writer(output, lineterminator="\n")
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
            csv_text = output.getvalue()
            safe_title = re.sub(r"[^A-Za-z0-9._-]+", "_", worksheet.title).strip("_")
            key = f"snapshots/{document_id}/{position}-{safe_title or 'sheet'}.csv"
            blob_store.put(key, csv_text.encode("utf-8"))
            combined.extend([f"Sheet: {worksheet.title}", csv_text, ""])
    finally:
        workbook.close()
    return _plain_text_pdf("\n".join(combined))


def _to_pdf(
    *,
    content: bytes,
    filename: str,
    mime: str,
    document_id: str,
    blob_store: BlobStore,
) -> bytes:
    suffix = Path(filename).suffix.casefold()
    if mime == "application/pdf" or suffix == ".pdf":
        return content
    if mime == "message/rfc822" or suffix == ".eml":
        return _plain_text_pdf(_email_text(content))
    if suffix == ".msg" or mime in {"application/vnd.ms-outlook", "application/x-msg"}:
        return _plain_text_pdf(_msg_text(content))
    if mime.startswith("image/") or suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
        return _image_pdf(content)
    if suffix == ".xlsx" or mime == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        return _xlsx_pdf_and_snapshots(content, document_id, blob_store)
    raise NormaliseError(f"unsupported document type: {mime or suffix}")


def _native_words(page: fitz.Page) -> tuple[list[dict[str, Any]], float]:
    width = page.rect.width
    height = page.rect.height
    page_area = width * height
    words = []
    covered = 0.0
    for raw in page.get_text("words"):
        x0, y0, x1, y1, text = raw[:5]
        covered += max(0.0, x1 - x0) * max(0.0, y1 - y0)
        words.append(
            {
                "text": text,
                "bbox": [x0 / width, y0 / height, x1 / width, y1 / height],
            }
        )
    return words, covered / page_area if page_area else 0.0


def normalise_document(
    *,
    document_id: str,
    filename: str,
    mime: str,
    source_key: str,
    blob_store: BlobStore,
    ocr_engine: OcrEngine | None,
) -> NormaliseResult:
    """Normalise one immutable original, render pages, and persist word streams."""

    try:
        source_bytes = blob_store.get(source_key)
        suffix = Path(filename).suffix.casefold()
        email_subject = None
        if mime == "message/rfc822" or suffix == ".eml":
            email_subject = _email_subject(source_bytes)
        elif suffix == ".msg" or mime in {
            "application/vnd.ms-outlook",
            "application/x-msg",
        }:
            email_subject = _msg_subject(source_bytes)
        pdf_bytes = _to_pdf(
            content=source_bytes,
            filename=filename,
            mime=mime,
            document_id=document_id,
            blob_store=blob_store,
        )
        pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
        if pdf.page_count < 1:
            raise NormaliseError("document has no readable pages")
        pdf_key = f"normalised/{document_id}.pdf"
        blob_store.put(pdf_key, pdf_bytes)
        text_keys = []
        page_keys = []
        first_text_parts = []
        page_text_coverages = []
        page_has_native_text = []
        try:
            for page_number, page in enumerate(pdf, start=1):
                pixmap = page.get_pixmap(matrix=fitz.Matrix(300 / 72, 300 / 72), alpha=False)
                png = pixmap.tobytes("png")
                page_key = f"pages/{document_id}/{page_number}.png"
                blob_store.put(page_key, png)
                page_keys.append(page_key)
                native_words, coverage = _native_words(page)
                page_text_coverages.append(coverage)
                page_has_native_text.append(bool(native_words))
                words = list(native_words)
                if coverage < 0.05 and ocr_engine is not None:
                    words.extend(ocr_engine.words(png))
                text_key = f"text/{document_id}/{page_number}.json"
                blob_store.put(
                    text_key,
                    json.dumps(words, separators=(",", ":"), ensure_ascii=False).encode("utf-8"),
                )
                text_keys.append(text_key)
                if len(" ".join(first_text_parts)) < 2_000:
                    first_text_parts.extend(str(word["text"]) for word in words)
        finally:
            pdf.close()
    except NormaliseError:
        raise
    except Exception as error:
        raise NormaliseError(str(error)) from error
    return NormaliseResult(
        pdf_key=pdf_key,
        page_count=len(page_keys),
        text_keys=text_keys,
        page_keys=page_keys,
        first_text=" ".join(first_text_parts)[:2_000],
        page_text_coverages=page_text_coverages,
        page_has_native_text=page_has_native_text,
        email_subject=email_subject,
    )
